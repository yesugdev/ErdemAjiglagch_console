import openpyxl
import openai
import requests
import os
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

class ExcelAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel өгөгдөлд дүн шинжилгээ хийх хэрэгсэл")
        self.root.geometry("1000x750")
        
        self.workbook = None
        self.current_sheet = None
        self.current_report = None
        self.report_type = None
        self.selected_sheet = None
        self.use_deepseek = tk.BooleanVar(value=True)  # Default to DeepSeek
        
        # Configure styles
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TLabel', background='#f0f0f0', font=('Arial', 10))
        self.style.configure('TButton', font=('Arial', 10))
        self.style.configure('Header.TLabel', font=('Arial', 12, 'bold'))
        
        # Create main container
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Header
        self.header = ttk.Label(self.main_frame, text="Excel өгөгдөлд дүн шинжилгээ хийх хэрэгсэл", style='Header.TLabel')
        self.header.pack(pady=(0, 20))
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs
        self.create_file_tab()
        self.create_sheet_tab()
        self.create_report_tab()
        self.create_settings_tab()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_bar = ttk.Label(self.main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        self.status_bar.pack(fill=tk.X, pady=(5, 0))
        self.update_status("Бэлэн")
        
    def update_status(self, message):
        self.status_var.set(message)
        self.root.update_idletasks()
        
    def create_file_tab(self):
        self.file_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.file_tab, text="1. Excel файл ачаалах")
        
        frame = ttk.Frame(self.file_tab)
        frame.pack(pady=20)
        
        ttk.Label(frame, text="Excel файлын зам:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.file_path_entry = ttk.Entry(frame, width=50)
        self.file_path_entry.grid(row=0, column=1, padx=5, pady=5)
        
        browse_btn = ttk.Button(frame, text="Файл сонгох", command=self.browse_file)
        browse_btn.grid(row=0, column=2, padx=5, pady=5)
        
        load_btn = ttk.Button(self.file_tab, text="Excel файл ачаалах", command=self.load_workbook)
        load_btn.pack(pady=10)
        
        self.sheet_info = tk.Text(self.file_tab, height=10, width=90, state=tk.DISABLED)
        self.sheet_info.pack(pady=10)
        
    def create_sheet_tab(self):
        self.sheet_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.sheet_tab, text="2. Хүснэгт сонгох", state=tk.DISABLED)
        
        frame = ttk.Frame(self.sheet_tab)
        frame.pack(pady=20)
        
        ttk.Label(frame, text="Хүснэгт сонгох:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.sheet_combobox = ttk.Combobox(frame, state="readonly")
        self.sheet_combobox.grid(row=0, column=1, padx=5, pady=5)
        self.sheet_combobox.bind("<<ComboboxSelected>>", self.on_sheet_select)
        
        ttk.Label(frame, text="Тайлангийн төрөл:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.report_type_combobox = ttk.Combobox(frame, 
                                               values=["SegmentCount", "PivotTable", "TimeResult", "segText"], 
                                               state="readonly")
        self.report_type_combobox.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(frame, text="Өгөгдлийн хүрээ (жишээ: A1:L5):").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.range_entry = ttk.Entry(frame)
        self.range_entry.grid(row=2, column=1, padx=5, pady=5)
        
        self.preview_btn = ttk.Button(frame, text="Өгөгдлийг үзэх", command=self.preview_data, state=tk.DISABLED)
        self.preview_btn.grid(row=3, column=0, columnspan=2, pady=10)
        
        self.data_preview = tk.Text(self.sheet_tab, height=10, width=90, state=tk.DISABLED)
        self.data_preview.pack(pady=10)
        
        self.generate_btn = ttk.Button(self.sheet_tab, text="Тайлан бэлтгэх", command=self.generate_report, state=tk.DISABLED)
        self.generate_btn.pack(pady=10)
        
    def create_report_tab(self):
        self.report_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.report_tab, text="3. Тайлан харах", state=tk.DISABLED)
        
        self.report_text = scrolledtext.ScrolledText(self.report_tab, wrap=tk.WORD, width=110, height=30, font=('Arial', 10))
        self.report_text.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
        
        btn_frame = ttk.Frame(self.report_tab)
        btn_frame.pack(pady=10)
        
        self.save_btn = ttk.Button(btn_frame, text="Тайлан хадгалах", command=self.save_report, state=tk.DISABLED)
        self.save_btn.pack(side=tk.LEFT, padx=5)
        
    def create_settings_tab(self):
        self.settings_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_tab, text="Тохиргоо")
        
        frame = ttk.Frame(self.settings_tab)
        frame.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="AI сервис сонгох:", font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5, pady=10, sticky=tk.W)
        
        deepseek_radio = ttk.Radiobutton(frame, text="DeepSeek AI", variable=self.use_deepseek, value=True)
        openai_radio = ttk.Radiobutton(frame, text="OpenAI", variable=self.use_deepseek, value=False)
        
        deepseek_radio.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        openai_radio.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        
        # API key sections
        ttk.Label(frame, text="DeepSeek API түлхүүр:", font=('Arial', 9)).grid(row=3, column=0, padx=5, pady=(20,5), sticky=tk.W)
        self.deepseek_key_entry = ttk.Entry(frame, width=50, show="*")
        self.deepseek_key_entry.grid(row=4, column=0, padx=5, pady=5, sticky=tk.W)
        
        ttk.Label(frame, text="OpenAI API түлхүүр:", font=('Arial', 9)).grid(row=5, column=0, padx=5, pady=(20,5), sticky=tk.W)
        self.openai_key_entry = ttk.Entry(frame, width=50, show="*")
        self.openai_key_entry.grid(row=6, column=0, padx=5, pady=5, sticky=tk.W)
        
        save_btn = ttk.Button(frame, text="Тохиргоог хадгалах", command=self.save_settings)
        save_btn.grid(row=7, column=0, pady=20, sticky=tk.W)
        
    def save_settings(self):
        # In a real application, you would securely save these API keys
        messagebox.showinfo("Амжилттай", "Тохиргоо хадгалагдлаа (жишээ програмын хувьд)")
        self.update_status("Тохиргоо шинэчлэгдлээ")
        
    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.file_path_entry.delete(0, tk.END)
            self.file_path_entry.insert(0, file_path)
            
    def load_workbook(self):
        file_path = self.file_path_entry.get()
        if not file_path:
            messagebox.showerror("Алдаа", "Эхлээд Excel файл сонгоно уу")
            return
            
        try:
            self.update_status("Файл ачаалж байна...")
            self.workbook = openpyxl.load_workbook(file_path)
            
            # Update sheet info
            self.sheet_info.config(state=tk.NORMAL)
            self.sheet_info.delete(1.0, tk.END)
            self.sheet_info.insert(tk.END, f"Амжилттай ачааллаа: {file_path}\n")
            self.sheet_info.insert(tk.END, f"\nБоломжит хүснэгтүүд: {', '.join(self.workbook.sheetnames)}")
            self.sheet_info.config(state=tk.DISABLED)
            
            # Update sheet combobox
            self.sheet_combobox['values'] = self.workbook.sheetnames
            self.notebook.tab(1, state=tk.NORMAL)  # Enable sheet tab
            self.update_status(f"Ачаалсан: {file_path}")
            
        except Exception as e:
            messagebox.showerror("Алдаа", f"Файл ачаалж чадсангүй: {e}")
            self.update_status(f"Алдаа: {str(e)}")
            
    def on_sheet_select(self, event):
        self.selected_sheet = self.sheet_combobox.get()
        if self.selected_sheet:
            if "SegmentCount" in self.selected_sheet:
                self.report_type_combobox.set("SegmentCount")
            elif "PivotTable" in self.selected_sheet:
                self.report_type_combobox.set("PivotTable")
            elif "TimeResult" in self.selected_sheet:
                self.report_type_combobox.set("TimeResult")
            elif "segText" in self.selected_sheet.lower():  # Case insensitive check
                self.report_type_combobox.set("segText")
            
            self.preview_btn.config(state=tk.NORMAL)
            
    def preview_data(self):
        cell_range = self.range_entry.get()
        if not cell_range:
            messagebox.showerror("Алдаа", "Өгөгдлийн хүрээ оруулна уу")
            return
            
        self.selected_sheet = self.sheet_combobox.get()
        self.report_type = self.report_type_combobox.get()
        
        if not self.selected_sheet or not self.report_type:
            messagebox.showerror("Алдаа", "Хүснэгт болон тайлангийн төрлийг сонгоно уу")
            return
            
        try:
            self.current_sheet = self.workbook[self.selected_sheet]
            data = self.get_data_range(cell_range)
            
            if data:
                self.data_preview.config(state=tk.NORMAL)
                self.data_preview.delete(1.0, tk.END)
                
                for row in data:
                    self.data_preview.insert(tk.END, "\t".join(row) + "\n")
                    
                self.data_preview.config(state=tk.DISABLED)
                self.generate_btn.config(state=tk.NORMAL)
                
        except Exception as e:
            messagebox.showerror("Алдаа", f"Өгөгдөл уншихад алдаа гарлаа: {e}")
            
    def get_data_range(self, cell_range):
        try:
            data = []
            for row in self.current_sheet[cell_range]:
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                data.append(row_data)
            return data
        except Exception as e:
            messagebox.showerror("Алдаа", f"Өгөгдөл авахад алдаа гарлаа: {e}")
            return None
            
    def generate_report(self):
        cell_range = self.range_entry.get()
        if not cell_range:
            messagebox.showerror("Алдаа", "Өгөгдлийн хүрээ оруулна уу")
            return
            
        data = self.get_data_range(cell_range)
        if not data:
            return
            
        self.update_status("Тайлан бэлтгэж байна...")
        
        system_prompt = {
            "role": "system", 
            "content": (
                "Та бол боловсролын судалгааны мэргэжилтэн. Та монгол хэл дээр хичээл судлаачид зориулсан мэргэжлийн "
                "дэлгэрэнгүй тайлан бэлтгэх ёстой. Тайлан үргэлж тоон утга дээр суурилсан байх ёстой. Тайлан нь дараах бүтэцтэй байх ёстой:\n"
                "1. Оршил: Судалгааны зорилго, ач холбогдол, судалгааны объект\n"
                "2. Арга зүй: Өгөгдөл цуглуулах, боловсруулах арга\n"
                "3. Дүн шинжилгээ: Нарийвчилсан дүн шинжилгээ, график, харьцуулалт\n"
                "4. Дүгнэлт: Гол дүгнэлт, санал, зөвлөмж\n"
                "Бүх хариултыг монгол хэлээр өгнө. Мэргэжлийн хэв маягтай, ойлгомжтой бичнэ үү."
            )
        }

        prompts = {
            "SegmentCount": (
                "Энэхүү өгөгдөл нь хэлэлцүүлгийн сегментүүдэд хуваагдсан. Сегмент бүрт асуултын тоо (Q), "
                "нээлттэй (QO), хаалттай (QC), бүтцийн (QL) асуултууд болон Блумын таксономын түвшин (Q1-Q6) "
                "бүртгэгдсэн. Дэлгэрэнгүй тайлан бэлтгэж өгнө үү. Тайлан нь:\n"
                "- Сегмент тус бүрийн асуултын төрлийн хуваарилалт\n"
                "- Блумын таксономын түвшний дүн шинжилгээ\n"
                "- Сургалтын үр нөлөөний талаарх дүгнэлт\n"
                "зэргийг агуулсан байх ёстой."
            ),
            "PivotTable": (
                "Энэхүү өгөгдлд сурагчид ба багш нарын хэрэглэсэн үгийн тоог харьцуулсан статистик байна. "
                "'ug' гэдэг нь 'үг' гэсэн үг. Дараах зүйлсийг тайлбарлана уу:\n"
                "- Багш, сурагчдын үгийн хэрэглээний харьцаа\n"
                "- Ялгаатай сегмент дэх үгийн хэрэглээний хэв маяг\n"
                "- Энэ хэрэглээ сургалтын үр нөлөөнд хэрхэн нөлөөлж болох талаарх дүгнэлт"
            ),
            "TimeResult": self._get_time_result_prompt(cell_range),
            "segText": (
                "Энэхүү өгөгдөл нь хэлэлцүүлгийн сегментүүдийн текстийн дүн шинжилгээний үр дүнг агуулсан. "
                "Хүснэгтийн багануудын тайлбар:\n"
                "- Segment: Хэлэлцүүлгийн сегментийг илэрхийлэх нэр. Энэ нь тухайн ярианы хэсгийг тодорхойлно.\n"
                "- Type: Ярилцагчийн төрөл:\n"
                "  - T: Багшийн яриа (Teacher)\n"
                "  - S: Сурагчийн яриа (Student)\n"
                "- Concatenated Values: Нэгдсэн эсвэл нийлмэл утга буюу текст.\n"
                "- Freq: word1, Freq: word2, Freq: word3 гэх мэт эдгээр багануудад тухайн харгалзах мөрөнд нь хэр их тус үгийг ашигласан давтамж орно.\n\n"
                "Дараах зүйлсийг тайлбарлана уу:\n"
                "- Багш, сурагчдын ярианы текстийн хэв маягийн ялгаа\n"
                "- Түлхүүр үгсийн давтамжийн дүн шинжилгээ\n"
                "- Сегмент тус бүр дэх хэлний хэрэглээний онцлог\n"
                "- Эдгээр үр дүнг сургалтын үр нөлөөтэй холбон дүгнэх"
            )
        }

        if self.report_type not in prompts:
            messagebox.showerror("Алдаа", "Тайлангийн төрөл буруу сонгогдсон")
            return
            
        data_str = "\n".join(["\t".join(row) for row in data])
        user_prompt = {
            "role": "user",
            "content": f"{prompts[self.report_type]}\n\nӨгөгдөл:\n{data_str}"
        }

        try:
            if self.use_deepseek.get():
                # DeepSeek API integration
                DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
                DEEPSEEK_API_KEY = self.deepseek_key_entry.get() or "YOUR_DEEPSEEK_API_KEY"
                
                headers = {
                    "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                    "Content-Type": "application/json"
                }
                
                payload = {
                    "model": "deepseek-chat",
                    "messages": [system_prompt, user_prompt],
                    "temperature": 0.7,
                    "max_tokens": 2000
                }
                
                response = requests.post(DEEPSEEK_API_URL, headers=headers, json=payload)
                response.raise_for_status()
                result = response.json()
                self.current_report = result['choices'][0]['message']['content']
                
            else:
                # OpenAI API integration
                openai.api_key = self.openai_key_entry.get() or "YOUR_OPENAI_API_KEY"
                response = openai.ChatCompletion.create(
                    model="gpt-4",
                    messages=[system_prompt, user_prompt],
                    temperature=0.7,
                    max_tokens=2500
                )
                self.current_report = response.choices[0].message.content
            
            # Display report
            self.report_text.config(state=tk.NORMAL)
            self.report_text.delete(1.0, tk.END)
            self.report_text.insert(tk.END, self.current_report)
            self.report_text.config(state=tk.DISABLED)
            
            # Enable report tab and save button
            self.notebook.tab(2, state=tk.NORMAL)
            self.save_btn.config(state=tk.NORMAL)
            self.notebook.select(2)  # Switch to report tab
            
            self.update_status("Тайлан амжилттай бэлтгэгдлээ")
            
        except Exception as e:
            messagebox.showerror("Алдаа", f"Тайлан бэлтгэхэд алдаа гарлаа: {e}")
            self.update_status(f"Алдаа: {str(e)}")
            
    def _get_time_result_prompt(self, cell_range):
        if cell_range and any(col in cell_range for col in ['F', 'G', 'H']):
            return (
                "Энэ хүснэгтэд багш болон сурагчдын ярих хурд (words/sec), ярих хугацаа (duration) "
                "зэргийг харуулсан. Дараах зүйлсийг тайлбарлана уу:\n"
                "- Ярианы хурдны дундаж утга, хэлбэлзэл\n"
                "- Багш, сурагчдын ярианы хугацааны харьцаа\n"
                "- Эдгээр үзүүлэлтүүд сургалтын чанарт хэрхэн нөлөөлж болох талаарх дүгнэлт"
            )
        elif cell_range and any(col in cell_range for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']):
            return (
                "Сегмент бүрт багш ба сурагчийн ярих хугацаа болон хурдыг харьцуулсан өгөгдөл. Дараах "
                "багануудыг анализ хийж тайлбарлана уу:\n"
                "- Segment: Хэлэлцүүлгийн сегмент\n"
                "- Teacher Time: Багшийн ярих хугацаа\n"
                "- Student Time: Сурагчийн ярих хугацаа\n"
                "- Teacher Words: Багшийн хэлсэн үгийн тоо\n"
                "- Student Words: Сурагчийн хэлсэн үгийн тоо\n"
                "- Teacher WPS: Багшийн үг/секунд\n"
                "- Student WPS: Сурагчийн үг/секунд\n"
                "- Total Duration: Нийт ярианы хугацаа\n"
                "- Teacher %: Багшийн эзлэх хувь\n"
                "- Student %: Сурагчийн эзлэх хувь\n"
                "Дүн шинжилгээнд сегмент тус бүрийн ялгаа, хэв маяг, сургалтын үр нөлөөний талаарх дүгнэлт орно."
            )
        else:
            return (
                "Энэ хүснэгтэд цаг хугацааны үр дүнг харуулсан. Дараах зүйлсийг тайлбарлана уу:\n"
                "- Ярианы хугацаа, хурдны үндсэн үзүүлэлтүүд\n"
                "- Багш, сурагчдын ярианы хэв маягийн ялгаа\n"
                "- Эдгээр үзүүлэлтүүд сургалтын үр дүнд хэрхэн нөлөөлж болох талаарх дүгнэлт"
            )
            
    def save_report(self):
        if not self.current_report:
            messagebox.showerror("Алдаа", "Хадгалах тайлан байхгүй байна")
            return
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"{self.report_type}_тайлан_{timestamp}.txt"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt")],
            initialfile=default_filename
        )
        
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(self.current_report)
                messagebox.showinfo("Амжилттай", f"Тайлан амжилттай хадгалагдлаа:\n{file_path}")
                self.update_status(f"Тайлан хадгалагдсан: {file_path}")
            except Exception as e:
                messagebox.showerror("Алдаа", f"Тайлан хадгалахад алдаа гарлаа: {e}")
                self.update_status(f"Тайлан хадгалах алдаа: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelAnalyzerGUI(root)
    root.mainloop()
