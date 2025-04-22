import openpyxl
import openai
import os
from datetime import datetime

class ExcelAnalyzer:
    def __init__(self):
        self.workbook = None
        self.current_sheet = None
        openai.api_key = "Key_oruulah"

    def load_workbook(self, file_path):
        """Excel файлыг ачаалах"""
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            print(f"\nАмжилттай: '{file_path}' файлыг ачааллаа.")
            print(f"Боломжит хуудаснууд: {', '.join(self.workbook.sheetnames)}")
            return True
        except Exception as e:
            print(f"\nАлдаа: Файл ачаалахад: {e}")
            return False

    def select_sheet_and_report_type(self):
        """Хуудас болон тайлангийн төрлийг нэг дор сонгох"""
        if not self.workbook:
            print("\nЭхлээд Excel файл ачаална уу!")
            return None, None
            
        print("\nХуудас болон тайлангийн төрөл сонгох:")
        sheets = self.workbook.sheetnames
        for i, sheet in enumerate(sheets, 1):
            print(f"{i}. {sheet}")
        
        try:
            choice = int(input("\nСонголтоо оруулна уу (1-3): "))
            if 1 <= choice <= len(sheets):
                selected_sheet = sheets[choice-1]
                self.current_sheet = self.workbook[selected_sheet]
                
                # Хуудасны нэрэнд тулгуурлан тайлангийн төрөл автоматаар тодорхойлох
                if "SegmentCount" in selected_sheet:
                    report_type = "SegmentCount"
                elif "PivotTable" in selected_sheet:
                    report_type = "PivotTable"
                elif "Time Result" in selected_sheet:
                    report_type = "TimeResult"
                else:
                    report_type = input("Тайлангийн төрөл оруулна уу (SegmentCount/PivotTable/TimeResult): ")
                
                print(f"\nСонгогдсон хуудас: {selected_sheet}")
                print(f"Тайлангийн төрөл: {report_type}")
                return selected_sheet, report_type
            else:
                print("\nБуруу сонголт! Дахин оролдоно уу.")
                return None, None
        except ValueError:
            print("\nТоон утга оруулна уу!")
            return None, None

    def get_data_range(self, cell_range):
        """Өгөгдлийн мужийг авах"""
        try:
            data = []
            for row in self.current_sheet[cell_range]:
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                data.append(row_data)
            return data
        except Exception as e:
            print(f"\nАлдаа: Өгөгдөл уншихад: {e}")
            return None

    def generate_report(self, data, report_type, cell_range=None):
        """Төрөл болон өгөгдөлд тулгуурлан тайлан бэлтгэх"""
        system_prompt = {
            "role": "system", 
            "content": (
                "Та бол боловсролын судалгааны мэргэжилтэн. Та монгол хэл дээр хичээл судлаачид зориулсан мэргэжлийн "
                "тайлан бэлтгэх ёстой. Тайлан нь дараах бүтэцтэй байх ёстой:\n"
                "1. Оршил: Судалгааны зорилго, ач холбогдол, судалгааны объект\n"
                "2. Арга зүй: Өгөгдөл цуглуулах, боловсруулах арга\n"
                "3. Дүн шинжилгээ: Нарийвчилсан дүн шинжилгээ, график, харьцуулалт\n"
                "4. Дүгнэлт: Гол дүгнэлт, санал, зөвлөмж\n"
                "Бүх хариултыг монгол хэлээр өгнө. Мэргэжлийн хэв маягтай, ойлгомжтой бичнэ үү."
            )
        }

        prompts = {
            "SegmentCount": (
                "Энэхүү өгөгдөл нь хэлэлцүүлгийн сегментүүдэд хуваагдсан. Сегмент бүрт асуултын тоо (Q), "
                "нээлттэй (QO), хаалттай (QC), бүтцийн (QL) асуултууд болон Блумын таксономын түвшин (Q1-Q6) "
                "бүртгэгдсэн. Дэлгэрэнгүй тайлан бэлтгэж өгнө үү. Тайлан нь:\n"
                "- Сегмент тус бүрийн асуултын төрлийн хуваарилалт\n"
                "- Блумын таксономын түвшний дүн шинжилгээ\n"
                "- Сургалтын үр нөлөөний талаарх дүгнэлт\n"
                "зэргийг агуулсан байх ёстой."
            ),
            "PivotTable": (
                "Энэхүү өгөгдөлд сурагчид ба багш нарын хэрэглэсэн үгийн тоог харьцуулсан статистик байна. "
                "'ug' гэдэг нь 'үг' гэсэн үг. Дараах зүйлсийг тайлбарлана уу:\n"
                "- Багш, сурагчдын үгийн хэрэглээний харьцаа\n"
                "- Ялгаатай сегмент дэх үгийн хэрэглээний хэв маяг\n"
                "- Энэ хэрэглээ сургалтын үр нөлөөнд хэрхэн нөлөөлж болох талаарх дүгнэлт"
            ),
            "TimeResult": self._get_time_result_prompt(cell_range)
        }

        if report_type not in prompts:
            print("\nАлдаа: Буруу тайлангийн төрөл!")
            return None

        data_str = "\n".join(["\t".join(row) for row in data])
        user_prompt = {
            "role": "user",
            "content": f"{prompts[report_type]}\n\nӨгөгдөл:\n{data_str}"
        }

        try:
            print("\nТайлан бэлтгэж байна...")
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[system_prompt, user_prompt],
                temperature=0.7,
                max_tokens=2500
            )
            return response.choices[0].message.content
        except Exception as e:
            print(f"\nАлдаа: ChatGPT API-д хандахад: {e}")
            return None

    def _get_time_result_prompt(self, cell_range):
        """TimeResult төрлийн prompt-ыг үүсгэх"""
        if cell_range and any(col in cell_range for col in ['F', 'G', 'H']):
            return (
                "Энэ хүснэгтэд багш болон сурагчдын ярих хурд (words/sec), ярих хугацаа (duration) "
                "зэргийг харуулсан. Дараах зүйлсийг тайлбарлана уу:\n"
                "- Ярианы хурдны дундаж утга, хэлбэлзэл\n"
                "- Багш, сурагчдын ярианы хугацааны харьцаа\n"
                "- Эдгээр үзүүлэлтүүд сургалтын чанарт хэрхэн нөлөөлж болох талаарх дүгнэлт"
            )
        elif cell_range and any(col in cell_range for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']):
            return (
                "Сегмент бүрт багш ба сурагчийн ярих хугацаа болон хурдыг харьцуулсан өгөгдөл. Дараах "
                "багануудыг анализ хийж тайлбарлана уу:\n"
                "- Segment: Хэлэлцүүлгийн сегмент\n"
                "- Teacher Time: Багшийн ярих хугацаа\n"
                "- Student Time: Сурагчийн ярих хугацаа\n"
                "- Teacher Words: Багшийн хэлсэн үгийн тоо\n"
                "- Student Words: Сурагчийн хэлсэн үгийн тоо\n"
                "- Teacher WPS: Багшийн үг/секунд\n"
                "- Student WPS: Сурагчийн үг/секунд\n"
                "- Total Duration: Нийт ярианы хугацаа\n"
                "- Teacher %: Багшийн эзлэх хувь\n"
                "- Student %: Сурагчийн эзлэх хувь\n"
                "Дүн шинжилгээнд сегмент тус бүрийн ялгаа, хэв маяг, сургалтын үр нөлөөний талаарх дүгнэлт орно."
            )
        else:
            return (
                "Энэ хүснэгтэд цаг хугацааны үр дүнг харуулсан. Дараах зүйлсийг тайлбарлана уу:\n"
                "- Ярианы хугацаа, хурдны үндсэн үзүүлэлтүүд\n"
                "- Багш, сурагчдын ярианы хэв маягийн ялгаа\n"
                "- Эдгээр үзүүлэлтүүд сургалтын үр дүнд хэрхэн нөлөөлж болох талаарх дүгнэлт"
            )

    def save_report(self, report, report_type):
        """Тайланг файлд хадгалах"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_тайлан_{timestamp}.txt"
        
        try:
            with open(filename, "w", encoding="utf-8") as f:
                f.write(report)
            print(f"\nТайлан '{filename}' файлд хадгалагдлаа.")
            return True
        except Exception as e:
            print(f"\nАлдаа: Файл хадгалахад: {e}")
            return False

def display_main_menu():
    """Үндсэн цэс харуулах"""
    print("\n===== Excel Өгөгдөл Анализийн Програм =====")
    print("1. Excel файл ачаалах")
    print("2. Хуудас сонгох, тайлангийн төрөл сонгох")
    print("3. Тайлан харах, хадгалах")
    print("4. Гарах")
    choice = input("\nСонголтоо оруулна уу (1-4): ")
    return choice

def main():
    analyzer = ExcelAnalyzer()
    current_report = None
    
    while True:
        choice = display_main_menu()
        
        if choice == "1":
            file_path = input("\nExcel файлын замыг оруулна уу: ")
            analyzer.load_workbook(file_path)
        
        elif choice == "2":
            selected_sheet, report_type = analyzer.select_sheet_and_report_type()
            if selected_sheet and report_type:
                cell_range = input("\nӨгөгдлийн мужийг оруулна уу (жишээ нь: A1:L5): ")
                data = analyzer.get_data_range(cell_range)
                
                if data:
                    current_report = analyzer.generate_report(
                        data, 
                        report_type,
                        cell_range if report_type == "TimeResult" else None
                    )
        
        elif choice == "3":
            if current_report:
                print("\n=== ТАЙЛАН ===\n")
                print(current_report)
                
                save_choice = input("\nТайланг хадгалах уу? (Y/N): ").lower()
                if save_choice == "y":
                    analyzer.save_report(current_report, report_type)
            else:
                print("\nЭхлээд тайлан бэлтгэнэ үү!")
        
        elif choice == "4":
            print("\nПрограмаас гарлаа. Баярлалаа!")
            break
        
        else:
            print("\nБуруу сонголт! 1-4 хооронд сонгоно уу.")

if __name__ == "__main__":
    main()
